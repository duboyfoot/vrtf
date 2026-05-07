"""
Mise en forme des feuilles de saisie du classeur BLD VRTF 1.1_modifiable.xlsx.
Applique couleurs, bordures et polices sans modifier les valeurs ni la structure.

Usage
-----
    py formater_saisie.py [--excel CHEMIN]
"""

import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

_HERE = Path(__file__).parent
_EXCEL_XLSM    = _HERE / "BLD VRTF 1.1_modifiable.xlsm"
_EXCEL_XLSX    = _HERE / "BLD VRTF 1.1_modifiable.xlsx"
_DEFAULT_EXCEL = _EXCEL_XLSM if _EXCEL_XLSM.exists() else _EXCEL_XLSX

# Palette identique à rapport.py
_DARK_BLUE  = "1F3864"
_MED_BLUE   = "2E75B6"
_PALE_BLUE  = "EBF3FB"
_LIGHT_BLUE = "BDD7EE"
_WHITE      = "FFFFFF"
_GRAY_LIGHT = "F2F2F2"
_GRAY_DARK  = "595959"
_YELLOW_IN  = "FFF2CC"   # cellules de saisie utilisateur


# ---------------------------------------------------------------------------
# Primitives de style
# ---------------------------------------------------------------------------

def _side(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt(cell, *, bold=False, italic=False, size=10, color="000000",
         bg=None, align="left", valign="center", border=False, wrap=False):
    cell.font = Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)
    if bg is not None:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border:
        cell.border = _border()


def _label(cell):
    _fmt(cell, bold=True, size=10, color=_DARK_BLUE, bg=_LIGHT_BLUE, border=True)


def _input(cell):
    _fmt(cell, size=10, bg=_YELLOW_IN, border=True, align="right")


def _unit(cell):
    _fmt(cell, size=9, italic=True, color=_GRAY_DARK, bg=_GRAY_LIGHT, border=True)


def _th(cell):
    _fmt(cell, bold=True, size=9, color=_WHITE, bg=_DARK_BLUE,
         align="center", border=True, wrap=True)


def _td(cell, alt=False):
    bg = _PALE_BLUE if alt else _WHITE
    align = "right" if isinstance(cell.value, (int, float)) else "center"
    _fmt(cell, size=9, bg=bg, border=True, align=align)


# ---------------------------------------------------------------------------
# Feuille Furnace design
# ---------------------------------------------------------------------------

# (col_index_0based_label, texte_label) → col_index_0based_valeur
_FD_PARAMS = {
    (0,  "Thichness :"):            2,
    (0,  "Width :"):                2,
    (0,  "Steel type :"):           2,
    (0,  "Production :"):           2,
    (0,  "entry strip T° :"):       2,
    (6,  "Wall thickness :"):       8,
    (6,  "Average conductivity :"): 8,
    (11, "Nombre Sections"):        13,
    (11, "Hauteur Section :"):      13,
    (11, "Largeur four:"):          13,
}

# Nombre de colonnes du tableau tubes (d'après lire_excel.py : indices 0-14)
_TUBE_N_COLS = 15


def _format_furnace_design(ws):
    # Largeurs de colonnes
    col_widths = {
        "A": 22, "B": 8,  "C": 14, "D": 10,
        "E": 8,  "F": 8,  "G": 24, "H": 8, "I": 14, "J": 10,
        "K": 8,  "L": 22, "M": 12, "N": 14, "O": 14,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    in_tube_table = False

    for row in ws.iter_rows():
        r = row[0].row
        row_vals = {cell.column - 1: cell.value for cell in row}

        # ── Paramètres scalaires ─────────────────────────────────────────────
        for (col_idx, text), val_col_idx in _FD_PARAMS.items():
            if row_vals.get(col_idx) == text:
                ws.row_dimensions[r].height = 18
                _label(ws.cell(row=r, column=col_idx + 1))
                _input(ws.cell(row=r, column=val_col_idx + 1))

        # ── Tableau tubes : en-tête ──────────────────────────────────────────
        if row_vals.get(0) == "Reelle zone":
            in_tube_table = True
            ws.row_dimensions[r].height = 30
            for c in range(1, _TUBE_N_COLS + 1):
                _th(ws.cell(row=r, column=c))
            continue

        # ── Tableau tubes : données ──────────────────────────────────────────
        if in_tube_table:
            v0 = row_vals.get(0)
            if v0 is not None and isinstance(v0, (int, float)):
                try:
                    alt = int(v0) % 2 == 1
                    ws.row_dimensions[r].height = 15
                    for c in range(1, _TUBE_N_COLS + 1):
                        _td(ws.cell(row=r, column=c), alt=alt)
                except (TypeError, ValueError):
                    in_tube_table = False
            else:
                in_tube_table = False

        # ── Tableau sections (col M = ID section 1-20, col N = nb rangées) ──
        m_val = row_vals.get(12)   # col M (index 12)
        n_val = row_vals.get(13)   # col N (index 13)
        if (m_val is not None and isinstance(m_val, (int, float)) and
                n_val is not None and isinstance(n_val, (int, float)) and
                1 <= int(m_val) <= 20):
            alt = int(m_val) % 2 == 1
            _td(ws.cell(row=r, column=13), alt=alt)   # ID section
            _input(ws.cell(row=r, column=14))          # nb rangées (éditable)


# ---------------------------------------------------------------------------
# Feuille Combustion
# ---------------------------------------------------------------------------

def _format_combustion(ws):
    ws.column_dimensions["A"].width = 28

    for row in ws.iter_rows():
        r = row[0].row
        label_val = row[0].value
        if label_val is None:
            continue

        ws.row_dimensions[r].height = 18

        # Col A : libellé
        _label(row[0])

        # Cols B+ : cellules de saisie (valeur présente)
        for cell in row[1:]:
            if cell.value is not None:
                _input(cell)


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Mise en forme des feuilles de saisie Excel VRTF"
    )
    parser.add_argument("--excel", default=str(_DEFAULT_EXCEL),
                        help="Classeur source .xlsx")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERREUR : fichier introuvable : {excel_path}")
        return

    print(f"Ouverture : {excel_path.name}")
    wb = openpyxl.load_workbook(excel_path, keep_vba=excel_path.suffix == ".xlsm")

    if "Furnace design" in wb.sheetnames:
        print("  Mise en forme : Furnace design...")
        _format_furnace_design(wb["Furnace design"])
    else:
        print("  AVERTISSEMENT : feuille 'Furnace design' introuvable")

    if "Combustion" in wb.sheetnames:
        print("  Mise en forme : Combustion...")
        _format_combustion(wb["Combustion"])
    else:
        print("  AVERTISSEMENT : feuille 'Combustion' introuvable")

    wb.save(excel_path)
    print(f"Enregistré : {excel_path.name}")
    print("Terminé.")


if __name__ == "__main__":
    main()
