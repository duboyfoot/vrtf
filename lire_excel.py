"""
Extrait les données du classeur BLD VRTF 1.1_modifiable.xlsx
et génère vrtf_reel.plf (fichier projet JSON complet).

Les lignes radex (feuille Mesh) sont intégrées dans le .plf
sous la clé "radex_lines" : run_vrtf.py les utilise directement
sans relancer Modray.
"""

import json
from pathlib import Path
import openpyxl

_XLSM = Path(r"C:\Users\patrick pro\Documents\kappa-dubois\developpement\BLD VRTF 1.1_modifiable.xlsm")
_XLSX = Path(r"C:\Users\patrick pro\Documents\kappa-dubois\developpement\BLD VRTF 1.1_modifiable.xlsx")
SOURCE = _XLSM if _XLSM.exists() else _XLSX
OUT_PLF = SOURCE.parent / "vrtf_reel.plf"


def load_furnace_design(ws):
    """Extrait géométrie + données tubes de la feuille Furnace design."""
    rows = list(ws.iter_rows(values_only=True))

    # ── Paramètres globaux (lus directement d'après position connue) ─────────
    # Repérage par valeurs connues (plus robuste qu'indices fixes)
    data = {}
    tubes_raw = []
    in_tubes = False

    for row in rows:
        non_none = [(i, v) for i, v in enumerate(row) if v is not None]
        if not non_none:
            continue
        vals = dict(non_none)

        # Plusieurs valeurs peuvent être sur la même ligne → if indépendants

        if row[0] == "Thichness :":
            data["strip_thickness_m"] = round(float(row[2]) / 1000.0, 6)
        if row[0] == "Width :":
            data["strip_width_m"] = float(row[2])
        if row[0] == "Steel type :":
            data["steel_grade"] = str(row[2])
        if row[6] == "Wall thickness :":
            data["wall_thickness_m"] = float(row[8])
        if row[6] == "Average conductivity :":
            data["wall_conductivity"] = float(row[8])
        if row[11] == "Nombre Sections":
            data["n_zones"] = int(row[13])
        if row[11] == "Hauteur Section :":
            data["zone_height_m"] = float(row[13])
        if row[11] == "Largeur four:":
            data["zone_width_m"] = float(row[13])
        if row[0] == "Production :":
            data["strip_flowrate_kgs"] = round(float(row[2]) * 1000.0 / 3600.0, 4)
        if row[0] == "entry strip T° :":
            data["T_strip_in_K"] = float(row[2]) + 273.0

        if row[0] == "Reelle zone":
            in_tubes = True
            continue

        if in_tubes and row[0] is not None:
            try:
                real_zone = int(row[0])
            except (TypeError, ValueError):
                in_tubes = False
                continue
            tubes_raw.append(row)

    # ── rows_per_zone : collecte depuis les lignes de sections ──────────────
    # Lignes avec Sections ID (int) et Columns Nber (int)
    sections_map = {}
    for row in rows:
        # Sections ID en col 12, Columns Nber en col 13
        if row[12] is not None and isinstance(row[12], (int, float)) and \
           row[13] is not None and isinstance(row[13], (int, float)):
            sec_id = int(row[12])
            n_cols = int(row[13])
            if 1 <= sec_id <= 20:
                sections_map[sec_id] = n_cols

    n_zones = data.get("n_zones", max(sections_map) if sections_map else 4)
    rows_per_zone = [sections_map.get(i, 4) for i in range(1, n_zones + 1)]
    data["rows_per_zone"] = rows_per_zone

    # gap_band_m ≈ 2 × distance_strip_tube (premier tube trouvé)
    # (la feuille ne stocke pas le gap explicitement)
    if tubes_raw:
        dist_st = float(tubes_raw[0][9]) if tubes_raw[0][9] is not None else 0.35
    else:
        dist_st = 0.35
    data["gap_band_m"] = [round(dist_st * 2, 4)] * n_zones

    # ── Tubes ────────────────────────────────────────────────────────────────
    # Colonnes : Reelle zone, Section, Rangee, Type, A, B, C, D,
    #            1er_abs, Distance_S-T, Pitch, N_tubes, T_tube[°C], Eff, Emis, S_tube
    tubes = []
    for row in tubes_raw:
        try:
            tubes.append({
                "real_zone":        int(row[0]),
                "zone":             int(row[1]),
                "row":              int(row[2]),
                "type":             str(row[3]).strip(),
                "a":                float(row[4]),
                "b":                float(row[5]),
                "c":                float(row[6]),
                "d":                float(row[7]) if row[7] is not None else 0.0,
                "prem_abs_m":       float(row[8]),
                "dist_band_tube_m": float(row[9]),
                "pitch_m":          float(row[10]),
                "n_tubes":          int(row[11]),
                "T_tube_K":         round(float(row[12]) + 273.0, 1),
                "tube_eff_pct":     round(float(row[13]) * 100.0, 1),
                "emissivity":       float(row[14]),
            })
        except (TypeError, ValueError, IndexError):
            continue

    data["tubes"] = tubes
    return data


def load_combustion(ws):
    """Extrait les zones de combustion de la feuille Combustion."""
    rows = {row[0]: row for row in ws.iter_rows(values_only=True) if row[0]}

    def _get(key, idx=1):
        r = rows.get(key)
        return r[idx] if r and r[idx] is not None else None

    nb_top = int(_get("Nb bruleurs top") or 0)

    zones = []
    for k in range(1, nb_top + 1):
        power = _get("Puissances top", k)
        if power is None or float(power) <= 0:
            continue

        excess = _get("ExcesAir top", k)
        eff_rege = _get("EffRegeAir top", k) or 0.0
        aspi_rege = _get("AspiRegeAir top", k) or 0.0
        tair = _get("Tair top", k) or 293.0
        tfuel = _get("Tfuel top", k) or 293.0

        # Composition combustible (top)
        n_gaz = int(_get("Combustible_Ngaz top", k) or 0)
        fuel_compo = {}
        for g in range(1, n_gaz + 1):
            name = _get(f"Combustible_Names top {g}", k)
            pct  = _get(f"Combustible_Compo top {g}", k)
            if name and pct:
                fuel_compo[str(name)] = float(pct)

        # Composition comburant
        n_air = int(_get("Comburant_Ngaz top", k) or 0)
        air_compo = {}
        for g in range(1, n_air + 1):
            name = _get(f"Comburant_Names top {g}", k)
            pct  = _get(f"Comburant_Compo top {g}", k)
            if name and pct:
                air_compo[str(name)] = float(pct)

        zones.append({
            "fuel":               fuel_compo or "Gaz naturel",
            "air":                air_compo  or "Air_sec",
            "excess_air_pct":     float(excess) if excess is not None else 10.0,
            "power_W":            float(power),
            "T_fuel_K":           float(tfuel),
            "T_air_K":            float(tair) if aspi_rege == 0.0 else 293.0,
            "regen_air_eff_pct":  float(eff_rege),
            "regen_air_aspi_pct": float(aspi_rege),
        })

    return zones


def load_mesh(ws):
    """Extrait les lignes radex de la feuille Mesh."""
    lines = []
    for row in ws.iter_rows(values_only=True):
        v = row[0]
        if v is not None:
            lines.append(str(v))
    # lines[0] = "160" (compte), lines[1:] = échanges
    return lines


def load_files(ws) -> dict:
    """
    Lit les chemins Thermette et Modray depuis la feuille Files.
    Retourne {"thermette_dir": Path, "modray_dir": Path}.
    """
    col_c = [row[2] for row in ws.iter_rows(values_only=True)]
    result = {}
    for i, val in enumerate(col_c):
        if val == "Thermette directory :" and i + 1 < len(col_c) and col_c[i + 1]:
            result["thermette_dir"] = Path(col_c[i + 1])
        elif val == "Modray directory :" and i + 1 < len(col_c) and col_c[i + 1]:
            result["modray_dir"] = Path(col_c[i + 1])
    return result


def main():
    print(f"Lecture : {SOURCE.name}")
    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    print("  Feuille Furnace design...")
    cfg = load_furnace_design(wb["Furnace design"])

    print("  Feuille Combustion...")
    zones_comb = load_combustion(wb["Combustion"])
    cfg["zones_combustion"] = zones_comb

    print("  Feuille Mesh (radex)...")
    radex = load_mesh(wb["Mesh"])
    cfg["radex_lines"] = radex

    # Résumé
    print(f"\n  n_zones       = {cfg.get('n_zones')}")
    print(f"  rows_per_zone = {cfg.get('rows_per_zone')}")
    print(f"  zone_height_m = {cfg.get('zone_height_m')} m")
    print(f"  zone_width_m  = {cfg.get('zone_width_m')} m")
    print(f"  strip         = {cfg.get('strip_thickness_m')*1000:.2f} mm × {cfg.get('strip_width_m')} m")
    print(f"  debit bande   = {cfg.get('strip_flowrate_kgs'):.2f} kg/s ({cfg.get('strip_flowrate_kgs',0)*3.6:.1f} t/h)")
    print(f"  tubes         = {len(cfg.get('tubes', []))} rangées")
    print(f"  comb. zones   = {len(zones_comb)}")
    for i, z in enumerate(zones_comb, 1):
        print(f"    Zone {i} : P={z['power_W']/1e6:.1f} MW  excès={z['excess_air_pct']:.0f}%  regen={z['regen_air_eff_pct']:.0f}%/aspi={z['regen_air_aspi_pct']:.0f}%")
    print(f"  radex_lines   = {len(radex)} (dont 1 ligne de compte)")

    with open(OUT_PLF, "w", encoding="utf-8") as fh:
        json.dump(cfg, fh, ensure_ascii=False, indent=2)
    print(f"\nFichier projet : {OUT_PLF}")


if __name__ == "__main__":
    main()
