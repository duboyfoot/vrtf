"""
Génération de la feuille Rapport (mise en forme commerciale) dans le classeur résultats.
Appelé depuis calculer.py après écriture des résultats bruts.
"""

from datetime import date

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ---------------------------------------------------------------------------
# Palette de couleurs
# ---------------------------------------------------------------------------

_DARK_BLUE  = "1F3864"
_MED_BLUE   = "2E75B6"
_LIGHT_BLUE = "BDD7EE"
_PALE_BLUE  = "EBF3FB"
_WHITE      = "FFFFFF"
_GRAY_LIGHT = "F2F2F2"
_GRAY_MED   = "D9D9D9"
_GRAY_DARK  = "595959"
_ORANGE     = "C55A11"


# ---------------------------------------------------------------------------
# Helpers de style
# ---------------------------------------------------------------------------

def _side(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)


def _border_thin():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _border_bottom():
    return Border(bottom=_side("medium", _DARK_BLUE))


def _cell(ws, row, col, value=None, *,
          bold=False, italic=False, size=10,
          color="000000", bg=None,
          align="left", valign="center",
          border=False, border_bottom=False,
          fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border:
        c.border = _border_thin()
    elif border_bottom:
        c.border = _border_bottom()
    if fmt:
        c.number_format = fmt
    return c


def _merge(ws, r, c1, c2, value, **kw):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    return _cell(ws, r, c1, value, **kw)


def _section_header(ws, r, title, c1=2, c2=8):
    ws.row_dimensions[r].height = 22
    _merge(ws, r, c1, c2, f"  {title}",
           bold=True, size=10, color=_WHITE, bg=_DARK_BLUE, align="left")
    return r + 1


def _col_headers(ws, r, headers, cols, bg=_MED_BLUE):
    ws.row_dimensions[r].height = 20
    for col, hdr in zip(cols, headers):
        _cell(ws, r, col, hdr,
              bold=True, size=9, color=_WHITE, bg=bg,
              align="center", border=True, wrap=True)
    return r + 1


def _data_row(ws, r, values, cols, alt=False, bold=False):
    bg = _PALE_BLUE if alt else _WHITE
    ws.row_dimensions[r].height = 16
    for col, val in zip(cols, values):
        v, fmt = (val if isinstance(val, tuple) else (val, None))
        _cell(ws, r, col, v,
              size=10, bg=bg, bold=bold,
              align=("right" if isinstance(v, (int, float)) else "left"),
              border=True, fmt=fmt)
    return r + 1


def _param_row(ws, r, label, value, unit="", alt=False):
    bg = _PALE_BLUE if alt else _WHITE
    ws.row_dimensions[r].height = 16
    _cell(ws, r, 2, label,  size=10, bg=bg, border=True)
    _cell(ws, r, 3, value,  size=10, bg=bg, border=True,
          align=("right" if isinstance(value, (int, float)) else "left"))
    _cell(ws, r, 4, unit,   size=10, bg=bg, border=True, color=_GRAY_DARK)
    return r + 1


# ---------------------------------------------------------------------------
# Création de la feuille Rapport
# ---------------------------------------------------------------------------

def create_rapport(wb, cfg: dict, pp: dict, comb_results: list) -> None:
    """
    Crée (ou recrée) la feuille 'Rapport' dans le classeur wb.
    Présente les résultats de manière professionnelle.
    """
    if "Rapport" in wb.sheetnames:
        del wb["Rapport"]
    ws = wb.create_sheet("Rapport", 0)

    # ── Largeurs de colonnes ─────────────────────────────────────────────────
    widths = {"A": 2, "B": 30, "C": 16, "D": 16,
              "E": 16, "F": 16, "G": 16, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Données calculées ────────────────────────────────────────────────────
    n_zones       = cfg["n_zones"]
    rows_per_zone = cfg["rows_per_zone"]
    tube_rows     = [(i + 1, j + 1)
                     for i in range(n_zones)
                     for j in range(rows_per_zone[i])]

    strip_temps   = pp["strip_temp_K"]
    T_out_K       = max(strip_temps.values()) if strip_temps else cfg["T_strip_in_K"]
    prod_th       = cfg["strip_flowrate_kgs"] * 3.6
    burner_kW     = sum(z["power_W"] for z in cfg.get("zones_combustion", [])) / 1000.0
    strip_pow_kW  = (pp["E_strip_out_W"] - pp["E_strip_in_W"]) / 1000.0
    wall_kW       = pp["flux_mur_four_W"] / 1000.0
    spec_MJt      = burner_kW * 3.6 / prod_th if prod_th else 0.0

    r = 1

    # ══════════════════════════════════════════════════════════════════════════
    # EN-TÊTE
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[r].height = 40
    _merge(ws, r, 2, 5, "KAPPA-DUBOIS",
           bold=True, size=18, color=_WHITE, bg=_DARK_BLUE, align="center")
    _merge(ws, r, 6, 8, f"Date : {date.today().strftime('%d/%m/%Y')}",
           size=9, color=_WHITE, bg=_DARK_BLUE, align="center")
    r += 1

    ws.row_dimensions[r].height = 28
    _merge(ws, r, 2, 8,
           "FOUR À TUBES RADIANTS VERTICAUX — Rapport de calcul thermique",
           bold=True, size=13, color=_WHITE, bg=_MED_BLUE, align="center")
    r += 1

    ws.row_dimensions[r].height = 16
    _merge(ws, r, 2, 8,
           f"Nuance : {cfg['steel_grade']}   |   "
           f"Épaisseur : {cfg['strip_thickness_m']*1000:.2f} mm   |   "
           f"Largeur : {cfg['strip_width_m']*1000:.0f} mm   |   "
           f"Production : {prod_th:.1f} t/h",
           italic=True, size=9, color=_GRAY_DARK, bg=_GRAY_LIGHT, align="center")
    r += 2

    # ══════════════════════════════════════════════════════════════════════════
    # PARAMÈTRES PROJET
    # ══════════════════════════════════════════════════════════════════════════
    r = _section_header(ws, r, "PARAMÈTRES DU PROJET")

    params = [
        ("Nuance d'acier",              cfg["steel_grade"],                          ""),
        ("Épaisseur bande",             round(cfg["strip_thickness_m"] * 1000, 2),   "mm"),
        ("Largeur bande",               round(cfg["strip_width_m"] * 1000, 0),       "mm"),
        ("Débit bande",                 round(prod_th, 1),                           "t/h"),
        ("Température entrée bande",    round(cfg["T_strip_in_K"] - 273.0, 0),       "°C"),
        ("Température sortie bande",    round(T_out_K - 273.0, 1),                   "°C"),
        ("Nombre de zones",             cfg["n_zones"],                              ""),
        ("Rangées par zone",            " / ".join(str(x) for x in rows_per_zone),  ""),
        ("Hauteur de zone",             cfg.get("zone_height_m", ""),               "m"),
        ("Largeur du four",             cfg.get("zone_width_m", ""),               "m"),
    ]
    for i, (lbl, val, unit) in enumerate(params):
        r = _param_row(ws, r, lbl, val, unit, alt=(i % 2 == 1))

    r += 1

    # ══════════════════════════════════════════════════════════════════════════
    # COMBUSTION
    # ══════════════════════════════════════════════════════════════════════════
    r = _section_header(ws, r, "COMBUSTION PAR ZONE")

    cols_c = [2, 3, 4, 5, 6, 7, 8]
    hdrs_c = ["Zone", "Puissance\n[kW]", "Débit gaz\n[Nm³/h]",
              "Débit air\n[Nm³/h]", "CO₂\n[%]", "H₂O\n[%]", "Excès air\n[%]"]
    r = _col_headers(ws, r, hdrs_c, cols_c)

    for i, cr in enumerate(comb_results):
        co2 = cr["flue_compo"].get("CO2", 0)
        h2o = cr["flue_compo"].get("H2O", 0)
        vals = [
            (f"Zone {i+1}", None),
            (round(cr["power_W"] / 1000, 1), "#,##0.0"),
            (round(cr["fuel_Nm3h"], 1),       "#,##0.0"),
            (round(cr["air_Nm3h"], 0),        "#,##0"),
            (round(co2, 1),                   "0.0"),
            (round(h2o, 1),                   "0.0"),
            (round(cr["excess_air_pct"], 0),  "0"),
        ]
        r = _data_row(ws, r, vals, cols_c, alt=(i % 2 == 1))

    r += 1

    # ══════════════════════════════════════════════════════════════════════════
    # RÉSULTATS PAR RANGÉE DE TUBES
    # ══════════════════════════════════════════════════════════════════════════
    r = _section_header(ws, r, "RÉSULTATS PAR RANGÉE DE TUBES")

    cols_t = [2, 3, 4, 5, 6, 7]
    hdrs_t = ["Section-Rangée", "Flux tube\n[kW]", "Débit gaz\n[Nm³/h]",
              "Débit fumées\n[Nm³/h]", "T bande\n[°C]", "T fumées\n[°C]"]
    r = _col_headers(ws, r, hdrs_t, cols_t)

    for idx, ij in enumerate(tube_rows):
        t_bande = pp["strip_temp_K"].get(ij)
        t_fum   = pp["T_fum_K"].get(ij, 273.0)
        vals = [
            (f"S{ij[0]}_R{ij[1]}", None),
            (round(pp["flux_tube_W"].get(ij, 0) / 1000, 2), "#,##0.00"),
            (round(pp["deb_gaz_Nm3h"].get(ij, 0), 2),        "#,##0.00"),
            (round(pp["deb_fum_Nm3h"].get(ij, 0), 2),        "#,##0.00"),
            (round(t_bande - 273.0, 1) if t_bande else "—",  "0.0"),
            (round(t_fum - 273.0, 1),                         "0.0"),
        ]
        r = _data_row(ws, r, vals, cols_t, alt=(idx % 2 == 1))

    # Ligne totaux
    ws.row_dimensions[r].height = 18
    total_vals = [
        ("TOTAL", None),
        (round(pp["flux_tube_four_W"] / 1000, 2),       "#,##0.00"),
        (round(sum(pp["deb_gaz_Nm3h"].values()), 2),    "#,##0.00"),
        (round(sum(pp["deb_fum_Nm3h"].values()), 2),    "#,##0.00"),
        ("", None),
        ("", None),
    ]
    bg_tot = _LIGHT_BLUE
    for col, val in zip(cols_t, total_vals):
        v, fmt = val
        _cell(ws, r, col, v,
              bold=True, size=10, bg=bg_tot,
              align=("right" if isinstance(v, (int, float)) else "center"),
              border=True, fmt=fmt)
    r += 2

    # ══════════════════════════════════════════════════════════════════════════
    # BILAN THERMIQUE
    # ══════════════════════════════════════════════════════════════════════════
    r = _section_header(ws, r, "BILAN THERMIQUE GLOBAL")

    bilan = [
        ("Production",                        round(prod_th, 1),                    "t/h"),
        ("Température entrée bande",          round(cfg["T_strip_in_K"] - 273, 0),  "°C"),
        ("Température sortie bande",          round(T_out_K - 273.0, 1),             "°C"),
        ("Puissance utile bande",             round(strip_pow_kW, 1),               "kW"),
        ("Flux tubes total",                  round(pp["flux_tube_four_W"]/1000, 1),"kW"),
        ("Pertes parois",                     round(wall_kW, 1),                    "kW"),
        ("Puissance brûleurs (installée)",    round(burner_kW, 1),                  "kW"),
        ("Efficacité thermique",
         round(strip_pow_kW / burner_kW * 100, 1) if burner_kW > 0 else 0.0,       "%"),
        ("Consommation spécifique",           round(spec_MJt, 1),                   "MJ/t"),
        ("Consommation spécifique",           round(spec_MJt / 4.1868, 1),          "kcal/kg"),
        ("Consommation spécifique",           round(spec_MJt / 3.6, 1),             "kWh/t"),
    ]

    for i, (lbl, val, unit) in enumerate(bilan):
        alt = (i % 2 == 1)
        bg  = _PALE_BLUE if alt else _WHITE
        ws.row_dimensions[r].height = 16
        bold_row = lbl.startswith("Efficacité") or lbl.startswith("Consommation spécifique")
        _cell(ws, r, 2, lbl,  size=10, bg=bg, border=True, bold=bold_row)
        _cell(ws, r, 3, val,  size=10, bg=bg, border=True, bold=bold_row,
              align="right",
              fmt=("#,##0.0" if isinstance(val, float) else None))
        _cell(ws, r, 4, unit, size=10, bg=bg, border=True, color=_GRAY_DARK)
        r += 1

    r += 2

    # ══════════════════════════════════════════════════════════════════════════
    # PIED DE PAGE
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[r].height = 14
    _merge(ws, r, 2, 8,
           "Document généré automatiquement par VRTF Python  —  "
           "github.com/duboyfoot/vrtf",
           italic=True, size=8, color=_GRAY_DARK, bg=_GRAY_LIGHT, align="center")

    # ── Paramètres d'impression ──────────────────────────────────────────────
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 9      # A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows = "1:3"
