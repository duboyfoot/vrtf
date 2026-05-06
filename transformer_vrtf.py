"""
Transforme BLD VRTF 1.1.xlsm :
  - Feuilles de données  → fichiers CSV dans le dossier 'data/'
  - Feuilles modifiables → nouveau classeur Excel 'BLD VRTF 1.1_modifiable.xlsx'
"""

import csv
import os
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SOURCE = Path(r"C:\Users\patrick pro\Documents\kappa-dubois\developpement\BLD VRTF 1.1.xlsm")
DATA_DIR = SOURCE.parent / "data"
OUTPUT_XLSX = SOURCE.parent / "BLD VRTF 1.1_modifiable.xlsx"

# Feuilles de données → CSV (lecture seule, référence)
DATA_SHEETS = {
    "basdo_gaz",
    "BisraCo",
    "BisraRo",
    "BisraCp",
    "Bisrah",
    "BisraEps",
    "Prop_comb",
    "Hottel",
    "Comburants",
    "Fuels",
    "ChaleurMassique",
    "Conductivité",
    "MasseVolumique",
    "Enthalpie",
    "Options",
    "texte",
}

# Feuilles modifiables → conservées dans le nouveau classeur Excel
MODIFIABLE_SHEETS = {
    "Furnace design",
    "Combustion",
    "Solver",
    "Mesh",
    "Recuperator",
    "Recuperator_old",
    "ResultsHEAT",
    "Results",
    "Files",
    "WData",
    "data",
}


def safe_filename(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_ " else "_" for c in name).strip()


def export_data_sheets(wb: openpyxl.Workbook) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    for sheet_name in wb.sheetnames:
        if sheet_name not in DATA_SHEETS:
            continue
        ws = wb[sheet_name]
        csv_path = DATA_DIR / f"{safe_filename(sheet_name)}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        print(f"  [DATA]  {sheet_name:25s} -> data/{csv_path.name}")


def copy_sheet_values(src_ws, dst_ws) -> None:
    """Copie valeurs + dimensions de colonnes/lignes (sans formules VBA)."""
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value

    for col_idx in range(1, src_ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in src_ws.column_dimensions:
            dst_ws.column_dimensions[col_letter].width = (
                src_ws.column_dimensions[col_letter].width
            )

    for row_idx, rd in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = rd.height


def export_modifiable_sheets(wb: openpyxl.Workbook) -> None:
    new_wb = Workbook()
    new_wb.remove(new_wb.active)  # retire la feuille vide par défaut

    for sheet_name in wb.sheetnames:
        if sheet_name not in MODIFIABLE_SHEETS:
            continue
        src_ws = wb[sheet_name]
        dst_ws = new_wb.create_sheet(title=sheet_name)
        copy_sheet_values(src_ws, dst_ws)
        print(f"  [XLSX]  {sheet_name:25s} -> {OUTPUT_XLSX.name}")

    new_wb.save(OUTPUT_XLSX)


def main() -> None:
    print(f"Lecture : {SOURCE.name}")
    wb = openpyxl.load_workbook(SOURCE, read_only=False, keep_vba=False, data_only=True)

    all_sheets = set(wb.sheetnames)
    unknown = all_sheets - DATA_SHEETS - MODIFIABLE_SHEETS
    if unknown:
        print(f"\nATTENTION — feuilles non classifiées (ignorées) : {unknown}\n")

    print("\n--- Export feuilles de données ---")
    export_data_sheets(wb)

    print("\n--- Export feuilles modifiables ---")
    export_modifiable_sheets(wb)

    print(f"\nTermine.")
    print(f"  Donnees  : {DATA_DIR}")
    print(f"  Classeur : {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
