"""
Workflow VRTF complet en une commande :
    1. Lecture du classeur Excel (Furnace design, Combustion, Mesh)
    2. Calcul combustion + Thermette
    3. Écriture des résultats dans les feuilles Results et ResultsHEAT

Usage
-----
    py calculer.py [--excel CHEMIN] [--thermette CHEMIN] [--out CHEMIN]

Options
-------
    --excel      Classeur source  (défaut : BLD VRTF 1.1_modifiable.xlsx dans ce dossier)
    --thermette  Exécutable       (défaut : C:\\thermette\\thermette.exe)
    --out        Classeur résultat (défaut : <nom>_résultats.xlsx à côté du source)
"""

import argparse
import io
import shutil
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

import vrtf
from lire_excel import load_furnace_design, load_combustion, load_mesh

_HERE = Path(__file__).parent
_DEFAULT_EXCEL = _HERE / "BLD VRTF 1.1_modifiable.xlsx"
_DEFAULT_THERM = Path(r"C:\thermette\thermette.exe")


# ---------------------------------------------------------------------------
# Écriture des résultats dans Excel
# ---------------------------------------------------------------------------

def _write_results(wb, cfg: dict, pp: dict, comb_results: list) -> None:
    """Remplit les feuilles Results et ResultsHEAT avec les résultats du calcul."""

    n_zones = cfg["n_zones"]
    rows_per_zone = cfg["rows_per_zone"]

    # Ordre (section, rangée) correspondant aux lignes 14-29 de la feuille Results
    tube_rows: list[tuple[int, int]] = [
        (i + 1, j + 1)
        for i in range(n_zones)
        for j in range(rows_per_zone[i])
    ]

    _write_results_sheet(wb["Results"], cfg, pp, comb_results, tube_rows)
    _write_results_heat_sheet(wb["ResultsHEAT"], cfg, pp, comb_results)


def _write_results_sheet(ws, cfg, pp, comb_results, tube_rows):
    """Feuille Results."""

    # ── C12 : pertes parois [kW] ──────────────────────────────────────────────
    ws["C12"] = round(pp["flux_mur_four_W"] / 1000.0, 2)

    # ── F14:F29 flux tube [kW], G14:G29 débit gaz [Nm³/h] ───────────────────
    for row_xl, ij in enumerate(tube_rows, start=14):
        ws[f"F{row_xl}"] = round(pp["flux_tube_W"].get(ij, 0.0) / 1000.0, 4)
        ws[f"G{row_xl}"] = round(pp["deb_gaz_Nm3h"].get(ij, 0.0), 4)

    # ── Synthèse (colonne L) ──────────────────────────────────────────────────
    prod_kgs   = cfg["strip_flowrate_kgs"]
    prod_th    = prod_kgs * 3.6                           # t/h
    T_in_C     = cfg["T_strip_in_K"] - 273.0
    burner_kW  = sum(z["power_W"] for z in cfg.get("zones_combustion", [])) / 1000.0

    strip_temps = pp["strip_temp_K"]
    T_out_K = max(strip_temps.values()) if strip_temps else cfg["T_strip_in_K"]
    T_out_C = T_out_K - 273.0

    strip_power_kW = (pp["E_strip_out_W"] - pp["E_strip_in_W"]) / 1000.0
    fume_flow      = sum(pp["deb_fum_Nm3h"].values())
    gas_flow       = sum(pp["deb_gaz_Nm3h"].values())
    wall_kW        = pp["flux_mur_four_W"] / 1000.0
    spec_cons      = burner_kW * 3.6 / prod_th           # MJ/t

    ws["L25"] = round(prod_th,        2)
    ws["L26"] = round(T_in_C,         1)
    ws["L27"] = round(T_out_C,        1)
    ws["L28"] = round(strip_power_kW, 2)
    ws["L29"] = round(fume_flow,      2)
    ws["L30"] = round(gas_flow,       2)
    ws["L31"] = round(wall_kW,        2)
    ws["L32"] = round(burner_kW,      2)
    ws["L33"] = round(spec_cons,      2)

    # ── Tableau BY SECTION (lignes 39-54) ─────────────────────────────────────
    for row_xl, ij in enumerate(tube_rows, start=39):
        q_tube = pp["flux_tube_W"].get(ij, 0.0)
        ws[f"K{row_xl}"] = round(q_tube / 1000.0, 4)
        ws[f"L{row_xl}"] = round(pp["deb_gaz_Nm3h"].get(ij, 0.0), 4)
        ws[f"N{row_xl}"] = round(pp["deb_fum_Nm3h"].get(ij, 0.0), 4)
        ws[f"O{row_xl}"] = round(pp["deb_fum_Nm3h"].get(ij, 0.0) / 3600.0 *
                                  1.2407, 5)                # ≈ kg/s (ρ fumées 0°C)
        ws[f"P{row_xl}"] = round(pp["T_fum_K"].get(ij, 273.0) - 273.0, 1)
        ws[f"Q{row_xl}"] = round((pp["T_fum_K"].get(ij, 273.0) - 273.0) * 9/5 + 32, 1)

    # ── Températures bande (section boundary) dans la colonne H ──────────────
    # Une valeur par ligne de tube : T en °C à l'entrée de la rangée suivante
    for row_xl, ij in enumerate(tube_rows, start=14):
        t_k = strip_temps.get(ij)
        if t_k is not None:
            ws[f"H{row_xl}"] = round(t_k - 273.0, 1)


def _write_results_heat_sheet(ws, cfg, pp, comb_results):
    """Feuille ResultsHEAT — bilan thermique global."""

    prod_kgs  = cfg["strip_flowrate_kgs"]
    prod_th   = prod_kgs * 3.6
    burner_W  = sum(z["power_W"] for z in cfg.get("zones_combustion", []))
    burner_kW = burner_W / 1000.0

    strip_in_kW  = pp["E_strip_in_W"]  / 1000.0
    strip_out_kW = pp["E_strip_out_W"] / 1000.0
    strip_pow_kW = strip_out_kW - strip_in_kW
    wall_kW      = pp["flux_mur_four_W"] / 1000.0
    spec_MJt     = burner_kW * 3.6 / prod_th
    spec_kcalkg  = spec_MJt / 4.1868              # MJ/t → kcal/kg  (1 MJ/t = 1/4.1868 kcal/kg)
    spec_kWht    = spec_MJt / 3.6                 # MJ/t → kWh/t

    # Entrées bilan
    ws["E36"] = round(strip_in_kW, 2)       # Puissance enthalpique bande entrée
    ws["E40"] = round(burner_kW,   2)       # Puissance combustion

    # Sorties bilan
    ws["J36"] = round(strip_out_kW, 2)      # Puissance enthalpique bande sortie
    ws["J38"] = round(wall_kW,      2)      # Pertes parois (murs)

    # Efficacité et consommations spécifiques
    if burner_kW > 0:
        ws["E26"] = round(strip_pow_kW / burner_kW * 100.0, 2)   # efficacité [%]
    ws["E28"] = round(spec_kcalkg, 2)
    ws["E29"] = round(spec_kWht,   2)
    ws["E30"] = round(spec_MJt,    2)

    # Données process
    ws["E8"]  = round(cfg["T_strip_in_K"]  - 273.0, 1)   # T entrée [°C]
    ws["E9"]  = round(cfg["strip_thickness_m"] * 1000.0, 2)  # épaisseur [mm]
    ws["E10"] = round(cfg["strip_width_m"]  * 1000.0, 1)     # largeur [mm]
    ws["E11"] = round(prod_kgs * 3.6, 2)                      # production [t/h]


# ---------------------------------------------------------------------------
# Pipeline complet
# ---------------------------------------------------------------------------

def run(excel_path: Path, thermette_exe: Path, out_path: Path) -> None:

    # 1. Lecture Excel
    print("=== Lecture Excel ===")
    wb_src = openpyxl.load_workbook(excel_path, data_only=True)
    cfg = load_furnace_design(wb_src["Furnace design"])
    cfg["zones_combustion"] = load_combustion(wb_src["Combustion"])
    radex_lines = load_mesh(wb_src["Mesh"])
    cfg["radex_lines"] = radex_lines
    print(f"  Acier  : {cfg['steel_grade']}")
    print(f"  Zones  : {cfg['n_zones']}  rangees/zone : {cfg['rows_per_zone']}")
    print(f"  Bande  : {cfg['strip_thickness_m']*1000:.2f} mm x {cfg['strip_width_m']} m")
    print(f"  Debit  : {cfg['strip_flowrate_kgs']:.2f} kg/s  ({cfg['strip_flowrate_kgs']*3.6:.1f} t/h)")
    print(f"  T_in   : {cfg['T_strip_in_K']-273:.0f} degC")
    print(f"  Radex  : {len(radex_lines)} lignes")

    # 2. Combustion
    print("\n=== Combustion ===")
    comb_results = vrtf.solve_combustion(cfg["zones_combustion"])
    for i, r in enumerate(comb_results, 1):
        print(f"  Zone {i} : P={r['power_W']/1e6:.2f} MW  gaz={r['fuel_Nm3h']:.1f} Nm3/h"
              f"  air={r['air_Nm3h']:.0f} Nm3/h")

    # 3. Fichiers Thermette
    workdir = excel_path.parent / "calcul"
    workdir.mkdir(parents=True, exist_ok=True)
    print("\n=== Génération fichiers Thermette ===")
    vrtf.write_thermette_files(workdir, cfg, radex_lines=radex_lines)
    print(f"  -> {workdir}")

    # 4. Solveur Thermette
    print("\n=== Solveur Thermette ===")
    if not thermette_exe.exists():
        raise FileNotFoundError(f"Thermette introuvable : {thermette_exe}")

    ther_dir = thermette_exe.parent
    for fname in ("VRTF_reseau", "VRTF_calc",
                  "AcierCp", "AcierCo", "AcierRo", "fluides.prp"):
        src = workdir / fname
        if src.exists():
            shutil.copy2(src, ther_dir / fname)
    print(f"  Exe    : {thermette_exe}")

    ret = vrtf.shell_wait(
        [str(thermette_exe), "VRTF_reseau", "VRTF_calc", "VRTF_resu"],
        cwd=ther_dir,
        timeout=120,
    )
    print(f"  Retour : {ret}")

    results_file = workdir / "VRTF_resu"
    resu_src = ther_dir / "VRTF_resu"
    if resu_src.exists():
        shutil.copy2(resu_src, results_file)
    else:
        raise RuntimeError("VRTF_resu absent apres calcul Thermette.")

    # 5. Post-traitement
    print("\n=== Post-traitement ===")
    z0 = cfg["zones_combustion"][0]
    pp = vrtf.postprocess(
        results_file, cfg,
        fuel_name=z0["fuel"],
        air_name=z0.get("air", "Air_sec"),
        excess_air_pct=z0.get("excess_air_pct", 0.0),
    )
    strip_temps = pp["strip_temp_K"]
    T_out_K = max(strip_temps.values()) if strip_temps else cfg["T_strip_in_K"]
    print(f"  T sortie bande     : {T_out_K-273:.1f} degC")
    print(f"  Puissance bande    : {(pp['E_strip_out_W']-pp['E_strip_in_W'])/1e3:.1f} kW")
    print(f"  Flux tubes total   : {pp['flux_tube_four_W']/1e3:.1f} kW")
    print(f"  Pertes parois      : {pp['flux_mur_four_W']/1e3:.1f} kW")
    burner_kW = sum(z["power_W"] for z in cfg["zones_combustion"]) / 1000.0
    prod_th = cfg["strip_flowrate_kgs"] * 3.6
    print(f"  Conso specifique   : {burner_kW*3.6/prod_th:.1f} MJ/t")

    # 6. Écriture Excel
    print("\n=== Écriture résultats Excel ===")
    wb_out = openpyxl.load_workbook(excel_path)   # avec formules pour conserver la mise en forme
    _write_results(wb_out, cfg, pp, comb_results)
    wb_out.save(out_path)
    print(f"  -> {out_path}")
    print("\nTermine.")


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Calcul VRTF : lecture Excel → Thermette → résultats dans Excel"
    )
    parser.add_argument("--excel",     default=str(_DEFAULT_EXCEL),
                        help="Classeur source .xlsx")
    parser.add_argument("--thermette", default=str(_DEFAULT_THERM),
                        help="Exécutable Thermette.exe")
    parser.add_argument("--out",       default=None,
                        help="Classeur résultat (défaut : <nom>_résultats.xlsx)")
    args = parser.parse_args()

    excel_path    = Path(args.excel)
    thermette_exe = Path(args.thermette)
    out_path      = (
        Path(args.out) if args.out
        else excel_path.parent / (excel_path.stem + "_résultats.xlsx")
    )

    if not excel_path.exists():
        print(f"ERREUR : classeur introuvable : {excel_path}", file=sys.stderr)
        sys.exit(1)

    run(excel_path, thermette_exe, out_path)


if __name__ == "__main__":
    main()
